# -*- coding: utf-8 -*-
from __future__ import print_function
from __future__ import division

from django.template.defaultfilters import slugify
from .models import ProductModel, Section

from openpyxl import load_workbook

import os
import six
import itertools


class ImportProductException(Exception):
    def __init__(self, reason, response=None):
        self.reason = six.text_type(reason)
        self.response = response
        Exception.__init__(self, reason)

    def __str__(self):
        return self.reason


class ImportProductModel(object):
    """docstring for ImportProductModel"""
    def __init__(self, filepath):
        super(ImportProductModel, self).__init__()
        self.filepath = filepath
        self.section_id = None
        self.code_list = None
        self.model_field = (
            ('A', 'section_id'),
            ('B', 'name'),
            ('C', 'search_name'),
            ('D', 'code'),
            ('E', 'description')
        )

    def is_valid_file_path(self):
        if not os.path.isfile(self.filepath):
            raise ImportProductException(
                "Incorrect path on file %s" % self.filepath)
        return True

    def get_all_products(self):
        if not self.section_id:
            raise ImportProductException(
                "Section_id not set")
        try:
            self.code_list = ProductModel.objects.filter(
                section_id=self.section_id).exclude(code='').values_list(
                    'code', flat=True)
        except Exception, e:
            raise ImportProductException(
                "Can't get information from database")

    def process(self):
        """
        Upload Product Model
        """
        self.is_valid_file_path()
        wb = load_workbook(self.filepath, use_iterators=True)
        if not wb.get_sheet_names():
            raise ImportProductException(
                "Sheet does not found")
        # Get firs sheet
        ws = wb.get_sheet_by_name(name=wb.get_sheet_names()[0])
        for row in ws.iter_rows():
            product_model = {}
            for cell in row:
                if cell.row == 1:
                    continue
                key = dict(self.model_field).get(cell.column)
                if not key:
                    continue
                if isinstance(cell.value, int):
                    product_model[key] = int(cell.value)
                else:
                    product_model[key] = cell.value.encode('utf-8')
            if product_model:
                # Upload all code in list
                if not self.section_id or \
                        self.section_id != product_model.get('section_id'):
                    self.section_id = product_model.get('section_id')
                    try:
                        Section.objects.get(pk=self.section_id)
                    except Section.DoesNotExist:
                        raise ImportProductException(
                            "Id %s not found" % self.section_id)
                    self.get_all_products()
                if product_model.get('code') not in self.code_list:
                    # product_model['inner_id'] = u''
                    product_model['alternative_connections'] = True
                    product_model['is_new'] = True
                    max_length = ProductModel._meta.get_field('slug').max_length
                    product_model['slug'] = orig = slugify(
                        product_model['name'])[:max_length]

                    for x in itertools.count(1):
                        if not ProductModel.objects.filter(
                                slug=orig).exists():
                            break
                        # Truncate the original slug dynamically. Minus 1 for the hyphen.
                        product_model['slug'] = "%s-%d" % (
                            orig[:max_length - len(str(x)) - 1], x)
                    try:
                        ProductModel.objects.create(**product_model)
                    except Exception:
                        raise ImportProductException(
                            "Can't insert %s" % product_model['name'])
